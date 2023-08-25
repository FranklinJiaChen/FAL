import datetime
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os
from anime_data_model import Anime

load_dotenv()
CLIENT_ID = os.getenv("CLIENT_ID")

def fetch_anime_data(anime_id):
    url = f'https://api.myanimelist.net/v2/anime/{anime_id}?fields=mean,num_favorites,statistics'
    response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
    response.raise_for_status()
    anime_data = response.json()
    print(anime_data)
    response.close()
    return anime_data

def create_anime(anime_id):
    anime_data = fetch_anime_data(anime_id)

    anime = Anime(anime_data)
    return anime

def create_sheet(workbook, originals_list):
    TITLE = 1
    FAVOURITE = 2
    PLAN_TO_WATCH = 3
    FAV_TO_PTW = 4

    headers = ['Title', 'Favourites', 'PTW', 'Favs:PTW']
    sheet = workbook.active

    for i, header in enumerate(headers):
        sheet.cell(1, i+1).value = header
        sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        sheet.cell(1, i+1).font = Font(bold=True)
        sheet.freeze_panes = 'B2'

    for j, anime in enumerate(originals_list):
        sheet.cell(j+2, TITLE).value = anime.title
        sheet.cell(j+2, FAVOURITE).value = anime.favourites
        sheet.cell(j+2, PLAN_TO_WATCH).value = anime.ptw
        sheet.cell(j+2, FAV_TO_PTW).value = anime.favourites_per_100_ptw()

        sheet.cell(j+2, TITLE).hyperlink = anime.get_mal_link()

    sheet[get_column_letter(PLAN_TO_WATCH) + '1'].comment = Comment("Plan to Watch", "Author")
    sheet[get_column_letter(FAV_TO_PTW) + '1'].comment = Comment("Favourite to Plan to Watch. Favourites/Plan to Watch * 100", "Author")

def main():
    ids = [54041, 55310, 54914, 55894, 53494, 53439, 44081, 50583, 54803, 54743, 52985, 54294, 55644, 51304, 54362, 54103, 53262, 54858, 54798, 53879, 53040, 54638, 52990, 54716, 54852, 54714, 55775, 52934, 54492, 55742, 53848, 50695, 50586, 51794, 53300, 54301, 51956, 54616, 51297, 44583, 50664, 51215, 50184, 52347, 53237, 52991, 53887, 52962, 42385, 54758, 54431, 54918, 52741, 49766, 53833, 55153]

    anime_list = [create_anime(anime_id) for anime_id in ids]

    workbook = Workbook()

    create_sheet(workbook, anime_list)

    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d-%H-%M")
    data_dir = "data"
    filename = os.path.join(data_dir, f"FAL_data_{timestamp}.xlsx")
    workbook.save(filename)

if __name__ == "__main__":
    main()
