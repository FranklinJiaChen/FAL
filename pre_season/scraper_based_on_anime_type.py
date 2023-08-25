import datetime
import requests
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os
from anime_data_model import Anime, AdaptedAnime, SequelAnime

load_dotenv()

CLIENT_ID = os.getenv("CLIENT_ID")

def fetch_anime_data(anime_id):
    url = f'https://api.myanimelist.net/v2/anime/{anime_id}?fields=mean,num_favorites,statistics,related_manga'
    response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
    response.raise_for_status()
    anime_data = response.json()
    print(anime_data)
    response.close()
    return anime_data

def fetch_manga_data(manga_id):
    url = f'https://api.myanimelist.net/v2/manga/{manga_id}?fields=mean,num_favorites,num_list_users,rank,media_type,related_manga'
    response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
    response.raise_for_status()
    manga_data = response.json()

    # print(manga_data["title"])
    # for related_manga in manga_data["related_manga"]:
    #     if related_manga["relation_type"] == "alternative_version":
    #         print((manga_data["id"], related_manga["node"]["id"]))
    response.close()
    return manga_data

def create_anime(anime_id):
    anime_data = fetch_anime_data(anime_id)

    anime = Anime(anime_data)
    return anime

def create_adapted_anime(anime_id, manga_id):
    anime_data = fetch_anime_data(anime_id)
    manga_data = fetch_manga_data(manga_id)

    adapted_anime = AdaptedAnime(
        anime_data, manga_data
    )
    return adapted_anime

def create_sequel_anime(anime_id, prequel_id, type, season):
    anime_data = fetch_anime_data(anime_id)
    prequel_data = fetch_anime_data(prequel_id)

    sequel_anime = SequelAnime(
        anime_data,
        prequel_data,
        type,
        season
    )
    return sequel_anime

def create_originals_sheet(workbook, originals_list):
    TITLE = 1
    FAVOURITE = 2
    PLAN_TO_WATCH = 3
    FAV_TO_PTW = 4

    headers = ['Title', 'Favourites', 'PTW', 'Favs:PTW']

    original_sheet = workbook.create_sheet(title="Originals")

    for i, header in enumerate(headers):
        original_sheet.cell(1, i+1).value = header
        original_sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        original_sheet.cell(1, i+1).font = Font(bold=True)
        original_sheet.freeze_panes = 'B2'

    for j, anime in enumerate(originals_list):
        original_sheet.cell(j+2, TITLE).value = anime.title
        original_sheet.cell(j+2, FAVOURITE).value = anime.favourites
        original_sheet.cell(j+2, PLAN_TO_WATCH).value = anime.ptw
        original_sheet.cell(j+2, FAV_TO_PTW).value = anime.favourites_per_100_ptw()

        original_sheet.cell(j+2, TITLE).hyperlink = anime.get_mal_link()

    original_sheet[get_column_letter(PLAN_TO_WATCH) + '1'].comment = Comment("Plan to Watch", "Author")
    original_sheet[get_column_letter(FAV_TO_PTW) + '1'].comment = Comment("Favourite to Plan to Watch. Favourites/Plan to Watch * 100", "Author")

def create_adaptations_sheet(workbook, adaptations_list):
    TITLE = 1
    FAVOURITE = 2
    PLAN_TO_WATCH = 3
    FAV_TO_PTW = 4
    TYPE = 5
    MANGA_NUM_USERS = 6
    MANGA_FAVOURITE = 7
    MANGA_SCORE = 8
    MANGA_RANK = 9
    MANGA_PERCENTILE = 10
    MANGA_FAV_TO_NUM_USERS = 11
    headers = ['Title', 'Favourites', 'PTW', 'Favs:PTW', 'TYPE', 'M_#Users', 'M_Favs', 'M_Score', 'M_Rank', 'M_%ile', 'M_Favs:Users']
    adaptation_sheet = workbook.create_sheet(title="Adaptations")

    for i, header in enumerate(headers):
        adaptation_sheet.cell(1, i+1).value = header
        adaptation_sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        adaptation_sheet.cell(1, i+1).font = Font(bold=True)
        adaptation_sheet.freeze_panes = 'B2'

    for j, anime in enumerate(adaptations_list):
        adaptation_sheet.cell(j+2, TITLE).value = anime.title
        adaptation_sheet.cell(j+2, FAVOURITE).value = anime.favourites
        adaptation_sheet.cell(j+2, PLAN_TO_WATCH).value = anime.ptw
        adaptation_sheet.cell(j+2, FAV_TO_PTW).value = anime.favourites_per_100_ptw()
        adaptation_sheet.cell(j+2, TYPE).value = anime.manga_type
        adaptation_sheet.cell(j+2, MANGA_NUM_USERS).value = anime.manga_num_list_users
        adaptation_sheet.cell(j+2, MANGA_FAVOURITE).value = anime.manga_favourite
        adaptation_sheet.cell(j+2, MANGA_SCORE).value = anime.manga_score
        adaptation_sheet.cell(j+2, MANGA_RANK).value = anime.manga_rank
        adaptation_sheet.cell(j+2, MANGA_PERCENTILE).value = anime.get_manga_percentile()
        adaptation_sheet.cell(j+2, MANGA_FAV_TO_NUM_USERS).value = anime.manga_favourite / anime.manga_num_list_users * 100


        adaptation_sheet.cell(j+2, TITLE).hyperlink = anime.get_mal_link()

    adaptation_sheet[get_column_letter(PLAN_TO_WATCH) + '1'].comment = Comment("Plan to Watch", "Author")
    adaptation_sheet[get_column_letter(FAV_TO_PTW) + '1'].comment = Comment("Favourite to Plan to Watch. Favourites/Plan to Watch * 100", "Author")
    adaptation_sheet[get_column_letter(MANGA_NUM_USERS) + '1'].comment = Comment("Number of users who have added the manga to their list", "Author")
    adaptation_sheet[get_column_letter(MANGA_FAVOURITE) + '1'].comment = Comment("Number of users who have favorited the manga", "Author")
    adaptation_sheet[get_column_letter(MANGA_SCORE) + '1'].comment = Comment("Mean score of the manga", "Author")
    adaptation_sheet[get_column_letter(MANGA_RANK) + '1'].comment = Comment("Rank of the manga", "Author")
    adaptation_sheet[get_column_letter(MANGA_PERCENTILE) + '1'].comment = Comment("Percentile of the manga", "Author")
    adaptation_sheet[get_column_letter(MANGA_FAV_TO_NUM_USERS) + '1'].comment = Comment("Favourites to Number of Users. Favourites/Number of Users * 100", "Author")

    return adaptation_sheet

def create_sequels_sheet(workbook, sequels_list):
    TITLE = 1
    FAVOURITE = 2
    PLAN_TO_WATCH = 3
    FAV_TO_PTW = 4
    TYPE = 5
    SEASON = 6
    PREQUEL_COMPLETED = 7
    PREQUEL_WATCHING = 8
    PREQUEL_DROPPED = 9
    PREQUEL_DROP_RATE = 10
    PREQUEL_RATING = 11

    headers = ['Title', 'Favourites', 'PTW', 'Favs:PTW', 'Type', 'Season', 'P Completed', 'P Watching', 'P Dropped', 'P Drop Rate', 'P Rating']
    sequel_sheet = workbook.create_sheet(title="Sequels")

    for i, header in enumerate(headers):
        sequel_sheet.cell(1, i+1).value = header
        sequel_sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        sequel_sheet.cell(1, i+1).font = Font(bold=True)
        sequel_sheet.freeze_panes = 'B2'

    for j, sequel_anime in enumerate(sequels_list):
        sequel_sheet.cell(j+2, TITLE).value = sequel_anime.title
        sequel_sheet.cell(j+2, FAVOURITE).value = sequel_anime.favourites
        sequel_sheet.cell(j+2, PLAN_TO_WATCH).value = sequel_anime.ptw
        sequel_sheet.cell(j+2, FAV_TO_PTW).value = sequel_anime.favourites_per_100_ptw()
        sequel_sheet.cell(j+2, TYPE).value = sequel_anime.sequel_type
        sequel_sheet.cell(j+2, SEASON).value = sequel_anime.season
        sequel_sheet.cell(j+2, PREQUEL_COMPLETED).value = sequel_anime.prequel.completed
        sequel_sheet.cell(j+2, PREQUEL_WATCHING).value = sequel_anime.prequel.watching
        sequel_sheet.cell(j+2, PREQUEL_DROPPED).value = sequel_anime.prequel.dropped
        sequel_sheet.cell(j+2, PREQUEL_DROP_RATE).value = sequel_anime.prequel.get_drop_rate()
        sequel_sheet.cell(j+2, PREQUEL_RATING).value = sequel_anime.prequel.rating

        sequel_sheet.cell(j+2, TITLE).hyperlink = sequel_anime.get_mal_link()

    sequel_sheet[get_column_letter(PLAN_TO_WATCH) + '1'].comment = Comment("Plan to Watch", "Author")
    sequel_sheet[get_column_letter(FAV_TO_PTW) + '1'].comment = Comment("Favourite to Plan to Watch. Favourites/Plan to Watch * 100", "Author")
    sequel_sheet[get_column_letter(PREQUEL_COMPLETED) + '1'].comment = Comment("Users Completed last part", "Author")
    sequel_sheet[get_column_letter(PREQUEL_WATCHING) + '1'].comment = Comment("Users Watching last part", "Author")
    sequel_sheet[get_column_letter(PREQUEL_DROPPED) + '1'].comment = Comment("Users Dropped last part", "Author")
    sequel_sheet[get_column_letter(PREQUEL_DROP_RATE) + '1'].comment = Comment("Of last part: Dropped / (Dropped + Completed + Watching) * 100", "Author")
    sequel_sheet[get_column_letter(PREQUEL_RATING) + '1'].comment = Comment("Rating of last part", "Author")


def main():
    # ids
    originals = [54914, 55894, 54798, 54638, 54301, 51956]
    # (id, manga_id)
    adaptations = [
    (54041, 134631), (55310, 124129), (53494, 115194), (53439, 113958), (44081, 98184), (50583, 142465), (52985, 124053),
    (54294, 155189), (54362, 120485), (54103, 143425), (53262, 127231), (53879, 130392), (52990, 144451), (54852, 132247),
    (54714, 123681), (50695, 107908), (52934, 130629), (54492, 110929), (53848, 145140), (50586, 107774), (53300, 119440),
    (54616, 110623), (51297, 106733), (50184, 124310), (52347, 130331), (53237, 121197), (52991, 126287), (52962, 121484),
    (54431, 74955), (52741, 123956), (49766, 114939), (53833, 129361), (55153, 115475)]
    # (id, prequel_id, sequel_type, season)
    sequels = [(54803, 36934, "sequel", 2), (54743, 53613, "part 2", 1.5), (55644, 48549, "part 2", 3.5), (51304, 227, "spin-off", 1), (54858, 40803, "sequel", 2), (53040, 43969, "sequel", 2),
                (54716, 3692, "far sequel", 2), (55775, 51139, "sequel", 2), (55742, 52955, "part 2", 2.5), (51794, 41491, "spin-off sequel", 2), (44583, 40958, "sequel", 2), (48761, 50664, "sequel", 2),
                  (51215, 42826, "sequel", 2), (53887, 50602, "sequel", 2), (42385, 10278, "spin-off", 1), (54758, 10278, "spin-off", 1), (54918, 50608, "sequel", 3)]

    #weird_prequel_adaptation = [(53848, 5711)]
    LNs = [(53494, 115193), (53439, 115165), (50583, 133333), (54103, 138763), (52990, 136419), (52934, 151180),
            (54492, 86769), (54616, 107569), (50184, 122521), (52962, 128337), (54431, 75121), (52741, 123960), (53833, 138424)]


    originals_list = [create_anime(anime_id) for anime_id in originals]
    adaptations_list = [create_adapted_anime(anime_id, manga_id) for anime_id, manga_id in adaptations]
    light_novels_list = [create_adapted_anime(anime_id, manga_id) for anime_id, manga_id in LNs]
    sequels_list = [create_sequel_anime(anime_id, prequel_id, type, season) for anime_id, prequel_id, type, season in sequels]

    workbook = Workbook()

    create_originals_sheet(workbook, originals_list)
    create_adaptations_sheet(workbook, adaptations_list)
    create_adaptations_sheet(workbook, light_novels_list)
    create_sequels_sheet(workbook, sequels_list)

    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d-%H-%M")
    filename = f"FAL_typed_data_{timestamp}.xlsx"
    workbook.save(filename)

if __name__ == "__main__":
    main()
