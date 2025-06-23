"""
scrapes every show in the upcoming season.
"""

import requests
from dotenv import load_dotenv
import os
from datetime import datetime
from anime_data_model import Anime
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

MAX_VALUE_FOR_LIMIT = 500
FIELDS = "mean, " \
"num_favorites, " \
"statistics, " \
"source, " \
"related_anime," \
"related_manga," \
"media_type"

load_dotenv()
CLIENT_ID = os.getenv("CLIENT_ID")

def get_upcoming_season() -> str:
    """
    Determines the upcoming Anime season based on the given date.

    Returns:
    str: The upcoming season (spring, summer, fall, winter) corresponding to the given date.
        February/March/April --> spring
        May/June/July --> summer
        August/September/October --> fall
        November/December/January --> winter
    """
    month = datetime.now().month

    if month in [2, 3, 4]:
        return 'spring'
    elif month in [5, 6, 7]:
        return 'summer'
    elif month in [8, 9, 10]:
        return 'fall'
    else:
        return 'winter'


def fetch_anime_data(anime_id: int, field: str) -> dict:
    """
    Fetches the data for a given anime from the MyAnimeList API.

    Parameters:
    anime_id: The ID of the anime to fetch.
    field: The fields to fetch for the anime.

    Returns:
    dict: The data for the anime from the MyAnimeList API.
    """
    url = f'https://api.myanimelist.net/v2/anime/{anime_id}?fields={field}'
    response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
    response.raise_for_status()
    anime_data = response.json()
    response.close()
    try:
        print(anime_data)
    except:
        print("can't print")
    return anime_data


def get_related_anime_id(anime: dict) -> tuple[int, str]:
    """
    Get the ID of the related anime.

    Parameters:
    anime: The anime data.

    Returns:
    tuple[int, str]: The ID of the related anime and the relation type.
    """
    try:
        return (anime['related_anime'][0]['node']['id'], anime['related_anime'][0]['relation_type'])
    except (KeyError, TypeError, IndexError):
        return (-1, "")

# def get_related_manga_id(anime: dict) -> tuple[int, str]:
#     """
#     [Does not work]
#     Get the ID of the related manga.

#     Parameters:
#     anime: The anime data.

#     Returns:
#     tuple[int, str]: The ID of the related manga and the relation type.
#     """
#     try:
#         return (anime['related_manga'][0]['node']['id'], anime['related_manga'][0]['relation_type'])
#     except (KeyError, TypeError, IndexError):
#         return (-1, "")

def fetch_anime_season_data(year: int, season: str, add_ids = [], unused_ids = []) -> [Anime, tuple[int, str]]:
    """
    Fetches the data for all the anime in a given season from the MyAnimeList API.

    Parameters:
    year: The year of the season.
    season: The season to fetch the anime for.

    Returns:
    [Anime, tuple[int, str]: The data for all the anime in
        the given season from the MyAnimeList API.
        where tuple[int, str] is the related anime id and the relation type.

    """
    url = f'https://api.myanimelist.net/v2/anime/season/{year}/{season}?limit={MAX_VALUE_FOR_LIMIT}'
    response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
    response.raise_for_status()
    anime_season_data = response.json()
    response.close()
    seasonal_list = [
        fetch_anime_data(anime['node']['id'], FIELDS)
        for anime in anime_season_data['data'] if anime['node']['id'] not in unused_ids
    ] + [fetch_anime_data(add_id, FIELDS) for add_id in add_ids]
    anime_list = [(Anime(anime), get_related_anime_id(anime)) for anime in seasonal_list if anime['media_type'] == 'tv']
    return anime_list


def create_sheet(workbook: str, anime_list: [Anime, tuple[int, str]]) -> None:
    """
    Creates a sheet in the workbook with the anime data.

    Parameters:
    workbook: The workbook to create the sheet in.
    anime_list: The list of anime data to add to the sheet.
    """
    TITLE = 1
    FAVOURITE = 2
    PLAN_TO_WATCH = 3
    FAV_TO_P2W = 4
    WATCHING = 5
    ID = 6
    SOURCE = 7
    PREQUEL_ID = 8
    RELATION_TO_PREQUEL = 9

    headers = ['Title', 'Favourites', 'P2W', 'Favs:P2W', 'Watching', 'id', 'Source', 'Prequel id', 'Relation2Prequel']
    sheet = workbook.active

    for i, header in enumerate(headers):
        sheet.cell(1, i+1).value = header
        sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        sheet.cell(1, i+1).font = Font(bold=True)
        sheet.freeze_panes = 'B2'

    for j, (anime, prequel) in enumerate(anime_list):
        sheet.cell(j+2, TITLE).value = anime.title
        sheet.cell(j+2, FAVOURITE).value = anime.favourites
        sheet.cell(j+2, PLAN_TO_WATCH).value = anime.p2w
        sheet.cell(j+2, FAV_TO_P2W).value = anime.favourites_per_100_p2w()
        sheet.cell(j+2, WATCHING).value = anime.watching
        sheet.cell(j+2, ID).value = anime.id
        sheet.cell(j+2, SOURCE).value = anime.source
        sheet.cell(j+2, PREQUEL_ID).value = prequel[0]
        sheet.cell(j+2, RELATION_TO_PREQUEL).value = prequel[1]

        sheet.cell(j+2, TITLE).hyperlink = anime.get_mal_link()
        sheet.cell(j+2, PREQUEL_ID).hyperlink = f"https://myanimelist.net/anime/{prequel[0]}"

    sheet[get_column_letter(PLAN_TO_WATCH) + '1'].comment = Comment("Plan to Watch", "Author")
    sheet[get_column_letter(FAV_TO_P2W) + '1'].comment = Comment("Favourite to Plan to Watch. Favourites/Plan to Watch * 100", "Author")


def main():
    """
    Fetches the data for all the anime in the upcoming season and prints it to the console.
    Creates an Excel file with the data.
    """
    # update these based on FAL season
    for_fal = 1 # boolean to add "fal" to the filename
    remove_ids = [
        53065,
        60543,
        59177,
        58390,
        57433,
        59986,
        60603,
        61793,
        59865,
        61731,
        61778
    ]

    add_ids = [
        52293,
        60326,
        61023,
        61150,
        57820,
        60505
    ]

    year = datetime.now().year
    upcoming_season = get_upcoming_season()
    anime_list = fetch_anime_season_data(year, upcoming_season, add_ids, remove_ids)
    for anime in anime_list:
        try:
            print(anime)
        except UnicodeEncodeError:
            print("Cannot print this anime due to encoding issues")
            pass

    workbook = Workbook()
    create_sheet(workbook, anime_list)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "data")
    filename = os.path.join(data_dir, f"{upcoming_season}{year}{'_fal'*for_fal}_shows.xlsx")
    workbook.save(filename)

if __name__ == "__main__":
    main()
