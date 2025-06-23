import datetime
import requests
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os
from anime_data_model import Anime, AdaptedAnime, SequelAnime

load_dotenv()
CLIENT_ID = os.getenv("CLIENT_ID")

def fetch_anime_data(anime_id):
    url = f'https://api.myanimelist.net/v2/anime/{anime_id}?fields=mean,num_favorites,statistics,related_manga,start_season'
    response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
    response.raise_for_status()
    anime_data = response.json()
    try:
        print(anime_data)
    except:
        print("cant print")
    response.close()
    return anime_data

def fetch_manga_data(manga_id):
    url = f'https://api.myanimelist.net/v2/manga/{manga_id}?fields=mean,num_favorites,num_list_users,rank,media_type,related_manga'
    response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
    response.raise_for_status()
    manga_data = response.json()

    # print(manga_data["title"])
    # for related_manga in manga_data["related_manga"]:
    #     if related_manga["relation_type"] == "alternative_version":
    #         print((manga_data["id"], related_manga["node"]["id"]))
    response.close()
    return manga_data

def create_anime(anime_id):
    anime_data = fetch_anime_data(anime_id)

    anime = Anime(anime_data)
    return anime

def create_adapted_anime(anime_id, manga_id):
    anime_data = fetch_anime_data(anime_id)
    manga_data = fetch_manga_data(manga_id)

    adapted_anime = AdaptedAnime(
        anime_data, manga_data
    )
    return adapted_anime

def create_sequel_anime(anime_id, prequel_id, type=-1, season="unknown"):
    anime_data = fetch_anime_data(anime_id)
    prequel_data = fetch_anime_data(prequel_id)

    sequel_anime = SequelAnime(
        anime_data,
        prequel_data,
        type,
        season
    )
    return sequel_anime

def create_originals_sheet(workbook, originals_list):
    TITLE = 1
    PLAN_TO_WATCH = 2
    FAVOURITE = 3
    FAV_TO_P2W = 4

    headers = ['Title', 'P2W', 'Favourites', 'Favs:P2W']

    original_sheet = workbook.create_sheet(title="Originals")

    for i, header in enumerate(headers):
        original_sheet.cell(1, i+1).value = header
        original_sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        original_sheet.cell(1, i+1).font = Font(bold=True)
        original_sheet.freeze_panes = 'B2'

    for j, anime in enumerate(originals_list):
        original_sheet.cell(j+2, TITLE).value = anime.title
        original_sheet.cell(j+2, PLAN_TO_WATCH).value = anime.p2w
        original_sheet.cell(j+2, FAVOURITE).value = anime.favourites
        original_sheet.cell(j+2, FAV_TO_P2W).value = anime.favourites_per_100_p2w()

        original_sheet.cell(j+2, TITLE).hyperlink = anime.get_mal_link()

    # Apply conditional formatting to the P2W column
    p2w_col_letter = get_column_letter(PLAN_TO_WATCH)
    p2w_range = f"{p2w_col_letter}2:{p2w_col_letter}{len(originals_list) + 1}"

    # note: Mettalic Rouge had 20150 and it flopped.
    # The only good orignal I have witnessed was Undead Girl Murder Farce which had 18235
    colour_rule = ColorScaleRule(
        start_type='num', start_value=5000, start_color='FF9999',    # Red
        mid_type='num', mid_value=10000, mid_color='FFFF99',      # Yellow
        end_type='num', end_value=25000, end_color='99FF99'       # Green
    )

    original_sheet.conditional_formatting.add(p2w_range, colour_rule)


    original_sheet[get_column_letter(PLAN_TO_WATCH) + '1'].comment = Comment("Plan to Watch", "Author")
    original_sheet[get_column_letter(FAV_TO_P2W) + '1'].comment = Comment("Favourite to Plan to Watch. Favourites/Plan to Watch * 100", "Author")

    original_sheet.column_dimensions[get_column_letter(FAVOURITE)].hidden = True
    original_sheet.column_dimensions[get_column_letter(FAV_TO_P2W)].hidden = True

def create_adaptations_sheet(workbook, adaptations_list):
    TITLE = 1
    FAVOURITE = 2
    PLAN_TO_WATCH = 3
    FAV_TO_P2W = 4
    TYPE = 5
    MANGA_NUM_USERS = 6
    MANGA_FAVOURITE = 7
    MANGA_SCORE = 8
    MANGA_RANK = 9
    MANGA_PERCENTILE = 10
    MANGA_FAV_TO_NUM_USERS = 11
    headers = ['Title', 'Favourites', 'P2W', 'Favs:P2W', 'TYPE', 'M_#Users', 'M_Favs', 'M_Score', 'M_Rank', 'M_%ile', 'M_Favs:Users']
    adaptation_sheet = workbook.create_sheet(title="Adaptations")

    for i, header in enumerate(headers):
        adaptation_sheet.cell(1, i+1).value = header
        adaptation_sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        adaptation_sheet.cell(1, i+1).font = Font(bold=True)
        adaptation_sheet.freeze_panes = 'B2'

    for j, anime in enumerate(adaptations_list):
        adaptation_sheet.cell(j+2, TITLE).value = anime.title
        adaptation_sheet.cell(j+2, FAVOURITE).value = anime.favourites
        adaptation_sheet.cell(j+2, PLAN_TO_WATCH).value = anime.p2w
        adaptation_sheet.cell(j+2, FAV_TO_P2W).value = anime.favourites_per_100_p2w()
        adaptation_sheet.cell(j+2, TYPE).value = anime.manga_type
        adaptation_sheet.cell(j+2, MANGA_NUM_USERS).value = anime.manga_num_list_users
        adaptation_sheet.cell(j+2, MANGA_FAVOURITE).value = anime.manga_favourite
        adaptation_sheet.cell(j+2, MANGA_SCORE).value = anime.manga_score
        adaptation_sheet.cell(j+2, MANGA_RANK).value = anime.manga_rank
        adaptation_sheet.cell(j+2, MANGA_PERCENTILE).value = anime.get_manga_percentile()
        adaptation_sheet.cell(j+2, MANGA_FAV_TO_NUM_USERS).value = anime.manga_favourite / anime.manga_num_list_users * 100


        adaptation_sheet.cell(j+2, TITLE).hyperlink = anime.get_mal_link()

    # Apply conditional formatting to the P2W column
    p2w_col_letter = get_column_letter(PLAN_TO_WATCH)
    p2w_range = f"{p2w_col_letter}2:{p2w_col_letter}{len(adaptations_list) + 1}"
    colour_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF9999',    # Red
        mid_type='num', mid_value=20000, mid_color='FFFF99',
        end_type='num', end_value=70000, end_color='99FF99'       # Green
    )
    adaptation_sheet.conditional_formatting.add(p2w_range, colour_rule)

    # Apply conditional formatting to the favourites column
    favourite_col_letter = get_column_letter(FAVOURITE)
    favourite_range = f"{favourite_col_letter}2:{favourite_col_letter}{len(adaptations_list) + 1}"
    favourite_colour_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF9999',    # Red
        mid_type='num', mid_value=100, mid_color='FFFF99',
        end_type='num', end_value=1000, end_color='99FF99'       # Green
    )
    adaptation_sheet.conditional_formatting.add(favourite_range, favourite_colour_rule)
    # Apply conditional formatting to the manga score column
    manga_score_col_letter = get_column_letter(MANGA_SCORE)
    manga_score_range = f"{manga_score_col_letter}2:{manga_score_col_letter}{len(adaptations_list) + 1}"
    manga_score_colour_rule = ColorScaleRule(
        start_type='num', start_value=6.5, start_color='FF9999',    # Red
        mid_type='num', mid_value=7.5, mid_color='FFFF99',
        end_type='num', end_value=8.5, end_color='99FF99'       # Green
    )
    adaptation_sheet.conditional_formatting.add(manga_score_range, manga_score_colour_rule)
    # Apply conditional formatting to the manga rank column
    # manga_rank_col_letter = get_column_letter(MANGA_RANK)
    # manga_rank_range = f"{manga_rank_col_letter}2:{manga_rank_col_letter}{len(adaptations_list) + 1}"
    # manga_rank_colour_rule = ColorScaleRule(
    #     start_type='num', start_value=10000, start_color='FF9999',    # Red
    #     mid_type='num', mid_value=5000, mid_color='FFFF99',
    #     end_type='num', end_value=1000, end_color='99FF99'       # Green
    # )
    # adaptation_sheet.conditional_formatting.add(manga_rank_range, manga_rank_colour_rule)

    # Apply conditional formatting to the manga percentile column
    manga_percentile_col_letter = get_column_letter(MANGA_PERCENTILE)
    manga_percentile_range = f"{manga_percentile_col_letter}2:{manga_percentile_col_letter}{len(adaptations_list) + 1}"
    manga_percentile_colour_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF9999',    # Red
        mid_type='num', mid_value=50, mid_color='FFFF99',
        end_type='num', end_value=100, end_color='99FF99'       # Green
    )
    adaptation_sheet.conditional_formatting.add(manga_percentile_range, manga_percentile_colour_rule)

    # Apply conditional format to manga num users column
    manga_num_users_col_letter = get_column_letter(MANGA_NUM_USERS)
    manga_num_users_range = f"{manga_num_users_col_letter}2:{manga_num_users_col_letter}{len(adaptations_list) + 1}"
    manga_num_users_colour_rule = ColorScaleRule(
        start_type='num', start_value=5000, start_color='FF9999',    # Red
        mid_type='num', mid_value=10000, mid_color='FFFF99',
        end_type='num', end_value=50000, end_color='99FF99'       # Green
    )
    adaptation_sheet.conditional_formatting.add(manga_num_users_range, manga_num_users_colour_rule)

    # Apply conditional format to manga favourites column
    manga_favourites_col_letter = get_column_letter(MANGA_FAVOURITE)
    manga_favourites_range = f"{manga_favourites_col_letter}2:{manga_favourites_col_letter}{len(adaptations_list) + 1}"
    manga_favourites_colour_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF9999',    # Red
        mid_type='num', mid_value=200, mid_color='FFFF99',
        end_type='num', end_value=2000, end_color='99FF99'       # Green
    )
    adaptation_sheet.conditional_formatting.add(manga_favourites_range, manga_favourites_colour_rule)

    adaptation_sheet[get_column_letter(PLAN_TO_WATCH) + '1'].comment = Comment("Plan to Watch", "Author")
    adaptation_sheet[get_column_letter(FAV_TO_P2W) + '1'].comment = Comment("Favourite to Plan to Watch. Favourites/Plan to Watch * 100", "Author")
    adaptation_sheet[get_column_letter(MANGA_NUM_USERS) + '1'].comment = Comment("Number of users who have added the manga to their list", "Author")
    adaptation_sheet[get_column_letter(MANGA_FAVOURITE) + '1'].comment = Comment("Number of users who have favorited the manga", "Author")
    adaptation_sheet[get_column_letter(MANGA_SCORE) + '1'].comment = Comment("Mean score of the manga", "Author")
    adaptation_sheet[get_column_letter(MANGA_RANK) + '1'].comment = Comment("Rank of the manga", "Author")
    adaptation_sheet[get_column_letter(MANGA_PERCENTILE) + '1'].comment = Comment("Percentile of the manga", "Author")
    adaptation_sheet[get_column_letter(MANGA_FAV_TO_NUM_USERS) + '1'].comment = Comment("Favourites to Number of Users. Favourites/Number of Users * 100", "Author")

    # hide ratio
    adaptation_sheet.column_dimensions[get_column_letter(FAV_TO_P2W)].hidden = True
    adaptation_sheet.column_dimensions[get_column_letter(MANGA_FAV_TO_NUM_USERS)].hidden = True
    adaptation_sheet.column_dimensions[get_column_letter(MANGA_PERCENTILE)].hidden = True
    adaptation_sheet.column_dimensions[get_column_letter(MANGA_RANK)].hidden = True


    return adaptation_sheet


def create_sequels_sheet(workbook, sequels_list):
    TITLE = 1
    FAVOURITE = 2
    PLAN_TO_WATCH = 3
    FAV_TO_P2W = 4
    TYPE = 5
    SEASON = 6
    PREQUEL_COMPLETED = 7
    PREQUEL_WATCHING = 8
    PREQUEL_DROPPED = 9
    PREQUEL_DROP_RATE = 10
    PREQUEL_RATING = 11
    PREQUEL_TITLE = 12
    PREQUEL_AIRING = 13

    headers = ['Title', 'Favourites', 'P2W', 'Favs:P2W', 'Type', 'Season', 'P Completed', 'P Watching', 'P Dropped', 'P Drop Rate', 'P Rating', 'P Title', "P air"]
    sequel_sheet = workbook.create_sheet(title="Sequels")

    for i, header in enumerate(headers):
        sequel_sheet.cell(1, i+1).value = header
        sequel_sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        sequel_sheet.cell(1, i+1).font = Font(bold=True)
        sequel_sheet.freeze_panes = 'B2'

    for j, sequel_anime in enumerate(sequels_list):
        sequel_sheet.cell(j+2, TITLE).value = sequel_anime.title
        sequel_sheet.cell(j+2, FAVOURITE).value = sequel_anime.favourites
        sequel_sheet.cell(j+2, PLAN_TO_WATCH).value = sequel_anime.p2w
        sequel_sheet.cell(j+2, FAV_TO_P2W).value = sequel_anime.favourites_per_100_p2w()
        sequel_sheet.cell(j+2, TYPE).value = sequel_anime.sequel_type
        sequel_sheet.cell(j+2, SEASON).value = sequel_anime.season
        sequel_sheet.cell(j+2, PREQUEL_COMPLETED).value = sequel_anime.prequel.completed
        sequel_sheet.cell(j+2, PREQUEL_WATCHING).value = sequel_anime.prequel.watching
        sequel_sheet.cell(j+2, PREQUEL_DROPPED).value = sequel_anime.prequel.dropped
        sequel_sheet.cell(j+2, PREQUEL_DROP_RATE).value = sequel_anime.prequel.get_drop_rate()
        sequel_sheet.cell(j+2, PREQUEL_RATING).value = sequel_anime.prequel.rating
        sequel_sheet.cell(j+2, PREQUEL_TITLE).value = sequel_anime.prequel.title
        sequel_sheet.cell(j+2, PREQUEL_AIRING).value = sequel_anime.prequel_airing

        sequel_sheet.cell(j+2, TITLE).hyperlink = sequel_anime.get_mal_link()
        sequel_sheet.cell(j+2, PREQUEL_TITLE).hyperlink = sequel_anime.prequel.get_mal_link()

    # Apply conditional formatting to the P2W column
    p2w_col_letter = get_column_letter(PLAN_TO_WATCH)
    p2w_range = f"{p2w_col_letter}2:{p2w_col_letter}{len(sequels_list) + 1}"
    colour_rule = ColorScaleRule(
        start_type='num', start_value=20000, start_color='FF9999',    # Red
        mid_type='num', mid_value=50000, mid_color='FFFF99',      # Yellow
        end_type='num', end_value=70000, end_color='99FF99'       # Green
    )
    sequel_sheet.conditional_formatting.add(p2w_range, colour_rule)
    # Apply conditional formatting to prequel completed column
    prequel_completed_col_letter = get_column_letter(PREQUEL_COMPLETED)
    prequel_completed_range = f"{prequel_completed_col_letter}2:{prequel_completed_col_letter}{len(sequels_list) + 1}"
    prequel_completed_colour_rule = ColorScaleRule(
        start_type='num', start_value=50000, start_color='FF9999',    # Red
        mid_type='num', mid_value=100000, mid_color='FFFF99',      # Yellow
        end_type='num', end_value=200000, end_color='99FF99'       # Green
    )
    sequel_sheet.conditional_formatting.add(prequel_completed_range, prequel_completed_colour_rule)
    # Apply conditional formatting to prequel rating column
    prequel_rating_col_letter = get_column_letter(PREQUEL_RATING)
    prequel_rating_range = f"{prequel_rating_col_letter}2:{prequel_rating_col_letter}{len(sequels_list) + 1}"
    prequel_rating_colour_rule = ColorScaleRule(
        start_type='num', start_value=6.5, start_color='FF9999',    # Red
        mid_type='num', mid_value=7.5, mid_color='FFFF99',      # Yellow
        end_type='num', end_value=8.5, end_color='99FF99'       # Green
    )
    sequel_sheet.conditional_formatting.add(prequel_rating_range, prequel_rating_colour_rule)

    # Add comments to the headers
    sequel_sheet[get_column_letter(PLAN_TO_WATCH) + '1'].comment = Comment("Plan to Watch", "Author")
    sequel_sheet[get_column_letter(FAV_TO_P2W) + '1'].comment = Comment("Favourite to Plan to Watch. Favourites/Plan to Watch * 100", "Author")
    sequel_sheet[get_column_letter(PREQUEL_COMPLETED) + '1'].comment = Comment("Users Completed last part", "Author")
    sequel_sheet[get_column_letter(PREQUEL_WATCHING) + '1'].comment = Comment("Users Watching last part", "Author")
    sequel_sheet[get_column_letter(PREQUEL_DROPPED) + '1'].comment = Comment("Users Dropped last part", "Author")
    sequel_sheet[get_column_letter(PREQUEL_DROP_RATE) + '1'].comment = Comment("Of last part: Dropped / (Dropped + Completed + Watching) * 100", "Author")
    sequel_sheet[get_column_letter(PREQUEL_RATING) + '1'].comment = Comment("Rating of last part", "Author")

    # hide
    sequel_sheet.column_dimensions[get_column_letter(FAV_TO_P2W)].hidden = True
    sequel_sheet.column_dimensions[get_column_letter(PREQUEL_DROP_RATE)].hidden = True
    sequel_sheet.column_dimensions[get_column_letter(PREQUEL_TITLE)].hidden = True
    sequel_sheet.column_dimensions[get_column_letter(PREQUEL_DROPPED)].hidden = True
    sequel_sheet.column_dimensions[get_column_letter(TYPE)].hidden = True
    sequel_sheet.column_dimensions[get_column_letter(SEASON)].hidden = True



def main():
    originals = [
        54028,
        61150,
        57820,
        58957,
        55689,
        60505
    ]
    adaptations = [
        (59961, 168633),
        (60732, 164607),
        (60951, 153011),
        (59791, 150111),
        (59459, 145359),
        (59161, 144290),
        (59845, 144267),
        (59062, 144180),
        (59207, 142649),
        (58913, 141833),
        (60260, 139086),
        (59424, 138928),
        (59205, 138396),
        (61023, 137439),
        (60315, 135486),
        (59689, 132750),
        (59421, 132404),
        (58197, 130317),
        (60131, 130181),
        (59130, 129759),
        (59619, 129418),
        (60326, 128538),
        (56907, 127971),
        (58811, 127436),
        (60523, 124973),
        (60130, 123921),
        (60508, 123519),
        (60697, 120467),
        (60665, 119119),
        (60316, 114601),
        (53512, 108492),
        (59632, 101946),
        (59898, 101554),
        (56693, 77827),
        (53397, 74283)
    ]

    sequels = [
        (61290, 59512),
        (60285, 58939),
        (61765, 58351),
        (61274, 57925),
        (61322, 57592),
        (54145, 54144),
        (61339, 53924),
        (58749, 53881),
        (59095, 53516),
        (59277, 53050),
        (56700, 52619),
        (59402, 48675),
        (58996, 42391),
        (57907, 40357),
        (61239, 27727),
        (52293, 8795),
        (59342, 2012)
    ]

    sequels_list = [create_sequel_anime(anime_id, prequel_id) for anime_id, prequel_id in sequels]
    originals_list = [create_anime(anime_id) for anime_id in originals]
    adaptations_list = [create_adapted_anime(anime_id, manga_id) for anime_id, manga_id in adaptations]
    # light_novels_list = [create_adapted_anime(anime_id, manga_id) for anime_id, manga_id in LNs]

    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    create_originals_sheet(workbook, originals_list)
    create_adaptations_sheet(workbook, adaptations_list)
    # create_adaptations_sheet(workbook, light_novels_list)
    create_sequels_sheet(workbook, sequels_list)

    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d-%H-%M")
    filename = f"FAL_typed_data_{timestamp}.xlsx"
    workbook.save(filename)

if __name__ == "__main__":
    main()
