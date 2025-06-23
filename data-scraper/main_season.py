"""
FAL season scraper

Current set for Spring 2024
"""

import datetime
import requests
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os
from anime_data_model import Anime

load_dotenv()
CLIENT_ID = os.getenv("CLIENT_ID")
FIELDS = "mean,num_favorites,statistics,source,related_anime,media_type"

PICKED_SHOWS = [
    "Bleach: Sennen Kessen-hen - Soukoku-tan",
    "Ao no Hako",
    "Dungeon ni Deai wo Motomeru no wa Machigatteiru Darou ka V: Houjou no Megami-hen",
    "Kimi wa Meido-sama.",
    "Kekkon suru tte, Hontou desu ka",
    "Arifureta Shokugyou de Sekai Saikyou Season 3",
    "Dragon Ball Daima",
    "Natsume Yuujinchou Shichi"
]

def fetch_anime_data(anime_id: int, field: str):
    url = f'https://api.myanimelist.net/v2/anime/{anime_id}?fields={field}'
    response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
    response.raise_for_status()
    anime_data = response.json()
    response.close()
    try:
        print(anime_data)
    except:
        print("can't print")
    return anime_data

def create_anime(anime_id):
    anime_data = fetch_anime_data(anime_id, FIELDS)

    anime = Anime(anime_data)
    return anime

def create_sheet(workbook, anime_list):
    TITLE = 1
    WATCHING = 2
    SCORE = 3
    DROPPED = 4
    FAVOURITE = 5
    P2W = 6
    DROP_RATE = 7
    COMPLETED = 8

    headers = ['Title', 'Watching', 'Score', 'Dropped', 'Favourite', 'P2W', 'Drop Rate', 'Completed']
    sheet = workbook.active

    for i, header in enumerate(headers):
        sheet.cell(1, i+1).value = header
        sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        sheet.cell(1, i+1).font = Font(bold=True)
        sheet.freeze_panes = 'B2'

    for j, anime in enumerate(anime_list):
        sheet.cell(j+2, TITLE).value = anime.title

        sheet.cell(j+2, WATCHING).value = anime.watching
        sheet.cell(j+2, SCORE).value = anime.rating
        sheet.cell(j+2, DROPPED).value = anime.dropped
        sheet.cell(j+2, FAVOURITE).value = anime.favourites
        sheet.cell(j+2, P2W).value = anime.p2w
        sheet.cell(j+2, DROP_RATE).value = anime.get_drop_rate()
        sheet.cell(j+2, COMPLETED).value = anime.completed

        sheet.cell(j+2, TITLE).hyperlink = anime.get_mal_link()
        sheet[get_column_letter(DROP_RATE) + '1'].comment = Comment("Dropped/(Dropped+Watching+Completed)", "Author")
        if anime.title in PICKED_SHOWS:
            sheet.cell(j+2, TITLE).font = Font(bold=True)

def main():
    # ids = [50869, 54233, 56768, 55102, 39894, 56738, 54309, 48418, 56165, 52196, 57623, 50713, 53410, 57798, 51859, 55855, 57946, 54758, 51122, 53407, 53835, 53516, 53434, 56923, 55265, 56690, 53912, 55717, 57100, 55877, 52588, 53770, 55597, 53865, 55911, 56838, 57031, 55194, 54859, 56980, 52405, 54900, 56230, 54199, 58080, 51461, 54839, 57391, 53356, 57614]
    ids = [
        59131,
        56967,
        56784,
        58516,
        57635,
        57066,
        58572,
        58511,
        57559,
        53287,
        55570,
        50306,
        52995,
        56400,
        55823,
        57102,
        56894,
        53033,
        57554,
        55994,
        57181,
        57611,
        55887,
        56228,
        56964,
        57891,
        55071,
        54853,
        55150,
        54726,
        57944,
        58445,
        57362,
        52215,
        56647,
        58714,
        56843,
        58172,
        57360,
        57796,
        57533,
        53723,
        56420,
        56662,
        58173,
        56461
    ]
    anime_list = [create_anime(anime_id) for anime_id in ids]

    workbook = Workbook()

    create_sheet(workbook, anime_list)

    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d-%H-%M")
    data_dir = "data"
    filename = os.path.join(data_dir, f"FAL_Fall_2024_data_{timestamp}.xlsx")
    workbook.save(filename)

if __name__ == "__main__":
    main()
